from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from nifty_stratlab.util.hashing import sha256_file


@dataclass
class WorkbookProfile:
    path: str
    sha256: str
    bytes: int
    sample_rows: int
    sheets: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        return "WARN" if self.warnings else "PASS"

    def as_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["status"] = self.status
        return payload


def _normalise(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def profile_workbook_structure(path: str | Path, *, sample_rows: int = 25) -> WorkbookProfile:
    """Inspect workbook structure and bounded samples without scanning all rows."""
    if sample_rows <= 0:
        raise ValueError("sample_rows must be positive")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to inspect workbook structure") from exc

    workbook_path = Path(path).expanduser().resolve()
    result = WorkbookProfile(
        path=str(workbook_path),
        sha256=sha256_file(workbook_path),
        bytes=workbook_path.stat().st_size,
        sample_rows=sample_rows,
    )
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, sample_rows), values_only=True))
        headers = [_normalise(value) for value in (rows[0] if rows else ()) if value is not None]
        sampled_dates = [
            value.date() if isinstance(value, datetime) else value
            for row in rows
            for value in row
            if isinstance(value, (date, datetime))
        ]
        result.sheets.append(
            {
                "name": worksheet.title,
                "reported_rows": worksheet.max_row,
                "reported_columns": worksheet.max_column,
                "sampled_rows": len(rows),
                "headers": headers,
                "first_sampled_date": min(sampled_dates).isoformat() if sampled_dates else None,
                "last_sampled_date": max(sampled_dates).isoformat() if sampled_dates else None,
            }
        )
    workbook.close()
    if not result.sheets:
        result.warnings.append("workbook contains no worksheets")
    result.warnings.append("structure/sample inspection only; workbook was not fully processed")
    result.warnings.append("available_at rule is unassigned; workbook is excluded from model features")
    return result
