#!/usr/bin/env python3
"""Export a governed 24-sheet strategy evaluation evidence pack."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import psycopg
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from psycopg.rows import dict_row

SHEETS = [
    "00 Executive Summary", "01 Rules & Identity", "02 Validation & Rankability",
    "03 Strategy Scorecard", "04 Intraday Target Ladder", "05 Swing Target Ladder",
    "06 Downside Ladder", "07 Capital Trap & Holding", "08 Monthly Performance",
    "09 Yearly Performance", "10 Regime Suitability", "11 Market-Stock Matrix",
    "12 Symbol Performance", "13 Sector Performance", "14 Time-of-Day",
    "15 Exit & Failure Analysis", "16 Portfolio & Capital", "17 Benchmark & Alpha",
    "18 No-Trade Rules", "19 P-Diagram Summary", "20 Data Quality", "21 Run Timing",
    "22 Manifest & Sources", "23 Trade Sample",
]


def fetch(cur, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur.execute(sql, params)
    return list(cur.fetchall())


def json_obj(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        return json.loads(value)
    return {}


def json_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    return []


def scalar(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def write_rows(ws, title: str, rows: Iterable[dict[str, Any]]) -> int:
    rows = list(rows)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="0B1F3A")
    if not rows:
        rows = [{"status": "NOT_ASSESSED", "note": "Canonical evidence is not available for this section."}]
    headers = list(rows[0])
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16624A")
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        ws.append([scalar(row.get(header)) for header in headers])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"
    for index, header in enumerate(headers, 1):
        width = max(12, min(42, len(str(header)) + 3))
        for cell in ws.iter_cols(min_col=index, max_col=index, min_row=3, max_row=min(ws.max_row, 80)):
            for item in cell:
                width = max(width, min(42, len(str(item.value or "")) + 2))
                item.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(index)].width = width
    return len(rows)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("status,note\nNOT_ASSESSED,No canonical rows available\n", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]), lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: scalar(value) for key, value in row.items()})


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--strategy-id", required=True)
    parser.add_argument("--scenario", default="nifty_100:capital_16l")
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    if not args.database_url:
        raise SystemExit("--database-url or DATABASE_URL is required")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    with psycopg.connect(args.database_url, row_factory=dict_row) as conn, conn.cursor() as cur:
        identity_rows = fetch(cur, """
          SELECT e.*,r.*,s.strategy_id,s.display_name,s.description,v.version_number,v.config_json,v.assumptions_json,p.policy_json
          FROM strategy_eval.v_latest_strategy_evaluation e
          JOIN nse_app.backtest_run r ON r.backtest_run_id=e.backtest_run_id
          JOIN nse_app.backtest_strategy_version v ON v.strategy_version_id=r.strategy_version_id
          JOIN nse_app.backtest_strategy s ON s.strategy_id=v.strategy_id
          JOIN strategy_eval.evaluation_policy p ON p.policy_version=e.policy_version
          WHERE s.strategy_id=%s AND r.scenario_key=%s LIMIT 1
        """, (args.strategy_id, args.scenario))
        if not identity_rows:
            raise SystemExit("No governed evaluation exists for that strategy/scenario")
        identity = identity_rows[0]
        evaluation_id = identity["evaluation_id"]
        run_id = identity["backtest_run_id"]
        summary = json_obj(identity["summary_json"])
        policy = json_obj(identity["policy_json"])
        validations = json_obj(identity["validation_json"])
        slices = fetch(cur, "SELECT slice_type,slice_key,sample_size,suitability,metrics_json FROM strategy_eval.slice_metric WHERE evaluation_id=%s ORDER BY slice_type,sample_size DESC", (evaluation_id,))
        stocks = fetch(cur, """SELECT symbol,security_name,sector,signal_count,accepted_trades,skipped_trades,win_rate_pct,avg_return_pct,median_return_pct,max_gain_pct,max_loss_pct,avg_hold_days,max_hold_days,realized_pnl,unrealized_pnl,charges,open_position_flag FROM nse_app.backtest_stock_summary WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s ORDER BY realized_pnl DESC NULLS LAST""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        trades = fetch(cur, """SELECT symbol,sector,signal_date,entry_date,exit_date,exit_reason,regime_on_entry,entry_price,exit_price,quantity,total_charges,net_pnl,profit_tax_amount,after_tax_net_pnl,return_pct,holding_days,trade_status FROM nse_app.backtest_trade_log WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s ORDER BY entry_date DESC,symbol LIMIT 1000""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        monthly = fetch(cur, """SELECT date_trunc('month',trade_date)::date period_start,(array_agg(total_equity ORDER BY trade_date))[1] start_equity,(array_agg(total_equity ORDER BY trade_date DESC))[1] end_equity,ROUND(100*((array_agg(total_equity ORDER BY trade_date DESC))[1]/NULLIF((array_agg(total_equity ORDER BY trade_date))[1],0)-1),4) return_pct FROM nse_app.backtest_daily_equity WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s GROUP BY 1 ORDER BY 1""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        yearly = fetch(cur, """SELECT date_trunc('year',trade_date)::date period_start,(array_agg(total_equity ORDER BY trade_date))[1] start_equity,(array_agg(total_equity ORDER BY trade_date DESC))[1] end_equity,ROUND(100*((array_agg(total_equity ORDER BY trade_date DESC))[1]/NULLIF((array_agg(total_equity ORDER BY trade_date))[1],0)-1),4) return_pct FROM nse_app.backtest_daily_equity WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s GROUP BY 1 ORDER BY 1""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        exits = fetch(cur, """SELECT COALESCE(exit_reason,'OPEN') exit_or_failure,COUNT(*) trade_count,ROUND(AVG(return_pct),4) avg_return_pct,ROUND(SUM(COALESCE(after_tax_net_pnl,net_pnl)),2) after_tax_pnl FROM nse_app.backtest_trade_log WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s GROUP BY 1 ORDER BY trade_count DESC""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        sectors = fetch(cur, """SELECT COALESCE(sector,'UNKNOWN') sector,COUNT(*) trade_count,ROUND(AVG(return_pct),4) avg_return_pct,ROUND(100*AVG(CASE WHEN COALESCE(after_tax_net_pnl,net_pnl)>0 THEN 1 ELSE 0 END),2) win_rate_pct,ROUND(SUM(COALESCE(after_tax_net_pnl,net_pnl)),2) after_tax_pnl FROM nse_app.backtest_trade_log WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s GROUP BY 1 ORDER BY after_tax_pnl DESC""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))
        sources = fetch(cur, "SELECT source_id,domain,source_type,quality_rank,source_url,usage_note FROM strategy_eval.source_register ORDER BY quality_rank,source_id")
        quality = fetch(cur, """SELECT 'market_events' metric,COUNT(*)::text value FROM strategy_eval.market_event UNION ALL SELECT 'point_in_time_eligible_events',COUNT(*) FILTER (WHERE point_in_time_eligible)::text FROM strategy_eval.market_event UNION ALL SELECT 'regime_rows',COUNT(*)::text FROM strategy_eval.market_regime_daily UNION ALL SELECT 'trade_context_rows',COUNT(*)::text FROM strategy_eval.trade_context_snapshot WHERE trade_log_id IN (SELECT trade_log_id FROM nse_app.backtest_trade_log WHERE batch_run_id=%s AND strategy_version_id=%s AND scenario_key=%s)""", (identity["batch_run_id"], identity["strategy_version_id"], identity["scenario_key"]))

    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheets = {name: workbook.create_sheet(name) for name in SHEETS}
    overview = [
        {"field": "Strategy", "value": identity["display_name"]}, {"field": "Scenario", "value": identity["scenario_label"]},
        {"field": "As of", "value": identity["as_of_date"]}, {"field": "Result type", "value": identity["result_type"]},
        {"field": "Rankability", "value": identity["rankability_status"]}, {"field": "Rating", "value": identity["rating"]},
        {"field": "Validation", "value": identity["validation_status"]}, {"field": "Verdict", "value": "Do not rank or promote" if identity["rankability_status"] != "RANKABLE" else "Eligible for governed comparison"},
    ]
    write_rows(worksheets[SHEETS[0]], SHEETS[0], overview)
    write_rows(worksheets[SHEETS[1]], SHEETS[1], [{"section": "config", "key": k, "value": v} for k, v in json_obj(identity["config_json"]).items()] + [{"section": "assumption", "key": k, "value": v} for k, v in json_obj(identity["assumptions_json"]).items()])
    write_rows(worksheets[SHEETS[2]], SHEETS[2], [{"gate": k, **v} for k, v in validations.items()])
    write_rows(worksheets[SHEETS[3]], SHEETS[3], [{"rating": identity["rating"], "quality_score": identity["quality_score"], "capacity_score": identity["revenue_capacity_score"], "note": "Scores remain blank while NOT_RANKABLE."}])
    write_rows(worksheets[SHEETS[4]], SHEETS[4], [{"target_pct": value, "evidence_status": "PATH_EVIDENCE_PENDING"} for value in policy["target_ladders_pct"]["intraday"]])
    write_rows(worksheets[SHEETS[5]], SHEETS[5], [{"target_pct": value, "deadline": "D+5", "evidence_status": "PATH_EVIDENCE_PENDING"} for value in policy["target_ladders_pct"]["swing_d1_to_d5"]])
    write_rows(worksheets[SHEETS[6]], SHEETS[6], [{"adverse_pct": value, "evidence_status": "MAE_EVIDENCE_PENDING"} for value in policy["adverse_ladder_pct"]])
    write_rows(worksheets[SHEETS[7]], SHEETS[7], [{"metric": key, "value": summary.get(key)} for key in ("avgHoldDays", "maxHoldDays", "openPositions", "maxOpenPositionsReached", "avgExposurePct", "exposurePct")])
    write_rows(worksheets[SHEETS[8]], SHEETS[8], monthly)
    write_rows(worksheets[SHEETS[9]], SHEETS[9], yearly)
    write_rows(worksheets[SHEETS[10]], SHEETS[10], slices)
    write_rows(worksheets[SHEETS[11]], SHEETS[11], [row for row in slices if row["slice_type"] == "STOCK_NIFTY_MATRIX"])
    write_rows(worksheets[SHEETS[12]], SHEETS[12], stocks)
    write_rows(worksheets[SHEETS[13]], SHEETS[13], sectors)
    write_rows(worksheets[SHEETS[14]], SHEETS[14], [{"status": "NOT_ASSESSED", "reason": "Published trade facts are daily and do not retain entry timestamps."}])
    write_rows(worksheets[SHEETS[15]], SHEETS[15], exits)
    write_rows(worksheets[SHEETS[16]], SHEETS[16], [{"metric": key, "value": value} for key, value in summary.items() if key in {"investedAmount", "currentValue", "cashBalance", "realizedPnl", "unrealizedPnl", "totalCharges", "taxDeducted", "totalReturnPct", "maxDrawdownPct"}])
    write_rows(worksheets[SHEETS[17]], SHEETS[17], [{"metric": key, "value": summary.get(key)} for key in ("benchmarkLabel", "benchmarkFinalValue", "excessOverBenchmark", "fdFinalValue", "excessOverFd")])
    write_rows(worksheets[SHEETS[18]], SHEETS[18], [{"rule": key, "status": value["status"], "rationale": value["reason"]} for key, value in validations.items() if value["status"] == "FAIL"])
    write_rows(worksheets[SHEETS[19]], SHEETS[19], [{"element": "Inputs", "treatment": "Versioned prices, regimes, VIX, point-in-time events"}, {"element": "Noise", "treatment": "Gaps, liquidity, costs, shocks, data uncertainty"}, {"element": "Controls", "treatment": "Hard blockers, capital limits, validation gates"}, {"element": "Outputs", "treatment": "NO_TRADE / WATCH / ELIGIBLE; never a broker order"}])
    write_rows(worksheets[SHEETS[20]], SHEETS[20], quality)
    write_rows(worksheets[SHEETS[21]], SHEETS[21], [{"run_id": run_id, "generated_at": identity["generated_at"], "evaluated_at": identity["evaluated_at"], "as_of_date": identity["as_of_date"]}])
    write_rows(worksheets[SHEETS[22]], SHEETS[22], sources)
    write_rows(worksheets[SHEETS[23]], SHEETS[23], trades)
    if monthly:
        chart = BarChart(); chart.title = "Monthly return (%)"; chart.y_axis.title = "Return %"; chart.x_axis.title = "Month"
        chart.add_data(Reference(worksheets[SHEETS[8]], min_col=4, min_row=2, max_row=worksheets[SHEETS[8]].max_row), titles_from_data=True)
        chart.set_categories(Reference(worksheets[SHEETS[8]], min_col=1, min_row=3, max_row=worksheets[SHEETS[8]].max_row))
        worksheets[SHEETS[8]].add_chart(chart, "F3")

    workbook_path = args.output_dir / "strategy_evaluation.xlsx"
    workbook.save(workbook_path)
    write_csv(args.output_dir / "trades.csv", trades)
    write_csv(args.output_dir / "slice_metrics.csv", slices)
    write_csv(args.output_dir / "stock_performance.csv", stocks)
    summary_json = {"strategy_id": args.strategy_id, "scenario": args.scenario, "evaluation_id": evaluation_id, "result_type": identity["result_type"], "rankability_status": identity["rankability_status"], "rating": identity["rating"], "validation_status": identity["validation_status"], "limitations": json_list(identity["limitation_json"]), "summary": summary}
    (args.output_dir / "strategy_summary.json").write_text(json.dumps(summary_json, indent=2, default=str), encoding="utf-8")
    (args.output_dir / "strategy_summary.md").write_text(
        f"# {identity['display_name']} — governed evaluation\n\n- Scenario: `{args.scenario}`\n- As of: {identity['as_of_date']}\n- Result type: **{identity['result_type']}**\n- Rankability: **{identity['rankability_status']}**\n- Rating: **{identity['rating']}**\n- Validation: **{identity['validation_status']}**\n\nThis run must not be promoted or ranked while mandatory validation gates fail. Retrospective event outcomes are analysis slices, not entry-time features.\n",
        encoding="utf-8",
    )
    artifacts = sorted(path for path in args.output_dir.iterdir() if path.is_file() and path.name != "checksums.sha256")
    (args.output_dir / "checksums.sha256").write_text("".join(f"{sha256(path)}  {path.name}\n" for path in artifacts), encoding="utf-8")
    artifacts = sorted(path for path in args.output_dir.iterdir() if path.is_file())
    with psycopg.connect(args.database_url) as manifest_conn, manifest_conn.cursor() as manifest_cur:
        for path in artifacts:
            manifest_cur.execute(
                """INSERT INTO strategy_eval.artifact_manifest (evaluation_id,artifact_type,artifact_path,sha256,size_bytes)
                   VALUES (%s,%s,%s,%s,%s)
                   ON CONFLICT (evaluation_id,artifact_path) DO UPDATE SET
                     artifact_type=EXCLUDED.artifact_type,sha256=EXCLUDED.sha256,size_bytes=EXCLUDED.size_bytes,created_at=NOW()""",
                (evaluation_id, path.suffix.lstrip(".") or "checksum", str(path.resolve()), sha256(path), path.stat().st_size),
            )
        manifest_conn.commit()
    print(json.dumps({"output_dir": str(args.output_dir), "files": len(artifacts), "sheets": len(SHEETS), "workbook_sha256": sha256(workbook_path)}, indent=2))


if __name__ == "__main__":
    main()
