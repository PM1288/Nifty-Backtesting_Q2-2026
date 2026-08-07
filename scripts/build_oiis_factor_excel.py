#!/usr/bin/env python3
"""Create the review workbook for completed O/X factor variations."""
from pathlib import Path
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SWEEP = ROOT / "platform/nifty_stratlab/outputs/oiis_49_factor_sweep_2026-08-07"
OUT = SWEEP / "OIIS_FACTOR_VARIATION_REVIEW.xlsx"

def main() -> None:
    ranking = pd.read_csv(SWEEP / "factor_ranking.csv")
    ranking["abs_p95_mae_pct"] = ranking["p95_mae_pct"].abs()
    ranking["opportunity_score"] = ranking["trade_count"] / ranking["trade_count"].max() * 100
    ranking["risk_score"] = (1 - ranking["abs_p95_mae_pct"] / ranking["abs_p95_mae_pct"].max()).clip(lower=0) * 100
    ranking["balanced_score"] = 0.60 * ranking["opportunity_score"] + 0.40 * ranking["risk_score"]
    ranking = ranking.sort_values(["balanced_score", "trade_count", "abs_p95_mae_pct"], ascending=[False, False, True])
    max_trades = ranking.trade_count.max(); least_mae = ranking.loc[ranking.trade_count == max_trades, "abs_p95_mae_pct"].min()
    best = ranking[(ranking.trade_count == max_trades) & (ranking.abs_p95_mae_pct == least_mae)]
    wb = Workbook(); readme = wb.active; readme.title = "00_READ_ME"
    readme.append(["OIIS O/X factor variation review"]); readme.append(["Scope", "40 completed of 49 requested variations; 9 require rerun after PostgreSQL DDL deadlock fix."])
    readme.append(["Best opportunity group", ", ".join(best.combination)])
    readme.append(["Trade count", int(max_trades)]); readme.append(["After-tax P&L", float(best.after_tax_pnl.iloc[0])]); readme.append(["P95 adverse excursion", float(best.p95_mae_pct.iloc[0])]); readme.append(["Selection rule", "Maximum executed trades, then least absolute P95 MAE; ties reported as a group."])
    for title, frame in (("01_RANKING", ranking), ("02_BEST_GROUP", best),):
        ws = wb.create_sheet(title); ws.append(list(frame.columns))
        for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1E3A5F")
        for row in frame.fillna("").itertuples(index=False, name=None): ws.append(list(row))
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    trade_rows=[]
    for combo in ranking.combination:
        matches=list(SWEEP.glob(f"{combo}/*/trades.csv"))
        if not matches: continue
        try: t=pd.read_csv(matches[0]); t.insert(0,"combination",combo); trade_rows.append(t)
        except pd.errors.EmptyDataError: pass
    if trade_rows:
        ws=wb.create_sheet("03_TRADE_LEDGER"); t=pd.concat(trade_rows,ignore_index=True); ws.append(list(t.columns))
        for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1E3A5F")
        for row in t.fillna("").itertuples(index=False,name=None): ws.append(list(row))
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    wb.save(OUT)
    (SWEEP / "FACTOR_REVIEW_SUMMARY.md").write_text(f"# OIIS factor variation review\n\nBest opportunity/risk group: {', '.join(best.combination)}\n\n- Trades: {int(max_trades)}\n- After-tax P&L: ₹{float(best.after_tax_pnl.iloc[0]):,.2f}\n- P95 MAE: {float(best.p95_mae_pct.iloc[0]):.4f}%\n- Selection: maximum trade count, then least absolute P95 adverse excursion.\n- This is a tie group, not a unique winner.\n- 40/49 variations completed; 9 failed during concurrent PostgreSQL schema DDL and must be rerun.\n")
    print(OUT)

if __name__ == "__main__": main()
