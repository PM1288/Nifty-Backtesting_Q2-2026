#!/usr/bin/env python3
"""Ingest the ROE workbook, calculate point-in-time regimes, and evaluate published runs.

This command is idempotent. It never treats retrospective event outcomes as an
entry-time feature unless the source row passes the strict point-in-time flag.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import psycopg
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MONOREPO_ROOT = PROJECT_ROOT.parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from nifty_stratlab.evaluation.roe import classify_trend, evaluate_rankability  # noqa: E402

POLICY_VERSION = "NIFTY-SEROE-V1.0"
EVALUATOR_VERSION = "roe-evaluator-1.0.0"
IST = ZoneInfo("Asia/Kolkata")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if np.isfinite(value) else None
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, np.generic):
        return jsonable(value.item())
    if isinstance(value, dict):
        return {str(k): jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [jsonable(v) for v in value]
    return str(value)


def as_json(value: Any) -> str:
    return json.dumps(jsonable(value), ensure_ascii=False, separators=(",", ":"))


def as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def as_ist_timestamp(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, (int, float)):
        parsed = from_excel(value)
    else:
        parsed = datetime.fromisoformat(str(value))
    return parsed.replace(tzinfo=IST) if parsed.tzinfo is None else parsed.astimezone(IST)


def sheet_records(workbook, sheet: str, header_row: int) -> list[dict[str, Any]]:
    ws = workbook[sheet]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row]]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        records.append(record)
    return records


def insert_policy(cur, policy_path: Path, rules_path: Path) -> dict[str, Any]:
    policy = json.loads(policy_path.read_text(encoding="utf-8"))
    cur.execute(
        """
        INSERT INTO strategy_eval.evaluation_policy
          (policy_version, policy_name, document_reference, effective_from, status, policy_json, source_sha256)
        VALUES (%s, %s, %s, %s, 'active', %s::jsonb, %s)
        ON CONFLICT (policy_version) DO UPDATE SET
          policy_name=EXCLUDED.policy_name, document_reference=EXCLUDED.document_reference,
          effective_from=EXCLUDED.effective_from, status='active', policy_json=EXCLUDED.policy_json,
          source_sha256=EXCLUDED.source_sha256
        """,
        (
            policy["policy_version"],
            policy["policy_name"],
            policy["document_reference"],
            policy["effective_from"],
            as_json(policy),
            sha256(rules_path),
        ),
    )
    return policy


def ingest_workbook(cur, workbook_path: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    required = {
        "00_README", "01_DASHBOARD", "02_EVENT_MASTER", "03_EVENT_WINDOWS",
        "04_DAILY_TEMPLATE", "05_CLASS_RULES", "06_DATA_DICTIONARY",
        "07_SOURCE_REGISTER", "08_LOOKUPS",
    }
    missing = sorted(required.difference(workbook.sheetnames))
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {missing}")
    workbook_hash = sha256(workbook_path)
    events = sheet_records(workbook, "02_EVENT_MASTER", 2)
    windows = sheet_records(workbook, "03_EVENT_WINDOWS", 2)
    sources = sheet_records(workbook, "07_SOURCE_REGISTER", 4)
    for record in events:
        event_timestamp = as_ist_timestamp(record.get("Event_Timestamp_IST"))
        point_in_time_eligible = bool(
            event_timestamp
            and record.get("Review_Status") == "Verified"
            and record.get("Data_Status") in {"calculated_continuous", "calculated_date_exception"}
            and not record.get("Overlap_Confounder")
        )
        cur.execute(
            """
            INSERT INTO strategy_eval.market_event (
              event_id, information_date, anchor_session, event_timestamp_ist, timestamp_precision,
              market_phase, event_name, event_category, event_subcategory, geography,
              scheduled_status, surprise_class, data_status, day_direction, day_magnitude,
              intraday_stress_zone, persistence_pattern, attribution_confidence, confidence_score,
              affected_sectors, overlap_confounder, review_status, point_in_time_eligible,
              raw_record, source_workbook_sha256)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s)
            ON CONFLICT (event_id) DO UPDATE SET
              information_date=EXCLUDED.information_date, anchor_session=EXCLUDED.anchor_session,
              event_timestamp_ist=EXCLUDED.event_timestamp_ist, timestamp_precision=EXCLUDED.timestamp_precision,
              market_phase=EXCLUDED.market_phase, event_name=EXCLUDED.event_name,
              event_category=EXCLUDED.event_category, event_subcategory=EXCLUDED.event_subcategory,
              geography=EXCLUDED.geography, scheduled_status=EXCLUDED.scheduled_status,
              surprise_class=EXCLUDED.surprise_class, data_status=EXCLUDED.data_status,
              day_direction=EXCLUDED.day_direction, day_magnitude=EXCLUDED.day_magnitude,
              intraday_stress_zone=EXCLUDED.intraday_stress_zone,
              persistence_pattern=EXCLUDED.persistence_pattern,
              attribution_confidence=EXCLUDED.attribution_confidence,
              confidence_score=EXCLUDED.confidence_score, affected_sectors=EXCLUDED.affected_sectors,
              overlap_confounder=EXCLUDED.overlap_confounder, review_status=EXCLUDED.review_status,
              point_in_time_eligible=EXCLUDED.point_in_time_eligible, raw_record=EXCLUDED.raw_record,
              source_workbook_sha256=EXCLUDED.source_workbook_sha256, ingested_at=NOW()
            """,
            (
                record["Master_Event_ID"], as_date(record["Information_Date"]), as_date(record["Anchor_Session"]),
                event_timestamp, record.get("Timestamp_Precision"), record.get("Market_Phase"), record["Event_Name"],
                record.get("Event_Category"), record.get("Event_Subcategory"), record.get("Geography"),
                record.get("Scheduled_Status"), record.get("Surprise_Class"), record.get("Data_Status"),
                record.get("Day_Direction"), record.get("Day_Magnitude"), record.get("Intraday_Stress_Zone"),
                record.get("Persistence_Pattern"), record.get("Attribution_Confidence"), record.get("Confidence_Score_0_100"),
                record.get("Affected_Sectors"), record.get("Overlap_Confounder"), record.get("Review_Status"),
                point_in_time_eligible, as_json(record), workbook_hash,
            ),
        )
    for record in windows:
        cur.execute(
            """
            INSERT INTO strategy_eval.event_window (
              event_id,horizon_label,sessions,window_start_date,window_end_date,return_pct,
              direction_class,magnitude_class,continuation_or_reversal,boundary_status,
              contamination_risk,raw_record)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)
            ON CONFLICT (event_id,horizon_label) DO UPDATE SET
              sessions=EXCLUDED.sessions,window_start_date=EXCLUDED.window_start_date,
              window_end_date=EXCLUDED.window_end_date,return_pct=EXCLUDED.return_pct,
              direction_class=EXCLUDED.direction_class,magnitude_class=EXCLUDED.magnitude_class,
              continuation_or_reversal=EXCLUDED.continuation_or_reversal,
              boundary_status=EXCLUDED.boundary_status,contamination_risk=EXCLUDED.contamination_risk,
              raw_record=EXCLUDED.raw_record
            """,
            (
                record["Master_Event_ID"], record["Horizon_Label"], int(record["Sessions"]),
                as_date(record.get("Window_Start_Date_Official")), as_date(record.get("Window_End_Date_Official")),
                None if record.get("Return_pct") is None else float(record["Return_pct"]) * 100.0,
                record.get("Direction_Class"), record.get("Magnitude_Class"), record.get("Continuation_or_Reversal"),
                record.get("Boundary_Status"), record.get("Contamination_Risk"), as_json(record),
            ),
        )
    for record in sources:
        cur.execute(
            """
            INSERT INTO strategy_eval.source_register
              (source_id,domain,source_type,quality_rank,source_url,usage_note,raw_record,source_workbook_sha256)
            VALUES (%s,%s,%s,%s,%s,%s,%s::jsonb,%s)
            ON CONFLICT (source_id) DO UPDATE SET domain=EXCLUDED.domain,source_type=EXCLUDED.source_type,
              quality_rank=EXCLUDED.quality_rank,source_url=EXCLUDED.source_url,usage_note=EXCLUDED.usage_note,
              raw_record=EXCLUDED.raw_record,source_workbook_sha256=EXCLUDED.source_workbook_sha256,ingested_at=NOW()
            """,
            (record["Source_ID"], record.get("Domain"), record.get("Source_Type"), record.get("Quality_Rank"),
             record.get("URL"), record.get("Usage_Note"), as_json(record), workbook_hash),
        )
    return {"events": len(events), "event_windows": len(windows), "sources": len(sources), "source_sheets": len(workbook.sheetnames)}


def fetch_frame(cur, sql: str, params: tuple[Any, ...] = ()) -> pd.DataFrame:
    cur.execute(sql, params)
    rows = cur.fetchall()
    return pd.DataFrame(rows, columns=[column.name for column in cur.description])


def _trend_series(values: pd.Series, thresholds: dict[str, float]) -> pd.Series:
    return values.map(lambda value: classify_trend(None if pd.isna(value) else float(value), **thresholds))


def calculate_regimes(cur, policy: dict[str, Any]) -> int:
    index_df = fetch_frame(
        cur,
        "SELECT trade_date,index_code AS symbol,close::double precision AS close_price FROM integration.v_index_daily_history ORDER BY index_code,trade_date",
    )
    cur.execute("SELECT batch_run_id FROM nse_app.batch_run_audit WHERE batch_name='backtesting_precompute' AND published_flag ORDER BY data_as_of_date DESC,generated_at DESC LIMIT 1")
    latest_batch = cur.fetchone()[0]
    stock_df = fetch_frame(
        cur,
        "SELECT trade_date,symbol,close_price::double precision FROM nse_app.backtest_feature_daily WHERE batch_run_id=%s ORDER BY symbol,trade_date",
        (latest_batch,),
    )
    index_df["instrument_type"] = "INDEX"
    stock_df["instrument_type"] = "STOCK"
    all_prices = pd.concat([index_df, stock_df], ignore_index=True)
    all_prices["trade_date"] = pd.to_datetime(all_prices["trade_date"])
    all_prices = all_prices.sort_values(["instrument_type", "symbol", "trade_date"])
    grouped = all_prices.groupby(["instrument_type", "symbol"], sort=False)["close_price"]
    all_prices["return_1d_pct"] = grouped.pct_change(1, fill_method=None) * 100
    all_prices["return_5d_pct"] = grouped.pct_change(5, fill_method=None) * 100
    all_prices["return_21d_pct"] = grouped.pct_change(21, fill_method=None) * 100
    all_prices["return_63d_pct"] = grouped.pct_change(63, fill_method=None) * 100
    all_prices["realised_vol_20d_pct"] = grouped.pct_change(fill_method=None).groupby(
        [all_prices["instrument_type"], all_prices["symbol"]]
    ).transform(lambda series: series.rolling(20, min_periods=20).std() * np.sqrt(252) * 100)
    thresholds = policy["trend_thresholds_pct"]
    for horizon in ("1d", "5d", "21d", "63d"):
        all_prices[f"trend_{horizon}"] = _trend_series(
            all_prices[f"return_{horizon}_pct"],
            {"bullish": thresholds[horizon]["bullish"], "bearish": thresholds[horizon]["bearish"], "sideways_abs": thresholds[horizon]["sideways_abs"]},
        )
    all_prices["primary_trend"] = all_prices.apply(
        lambda row: next((row[f"trend_{h}"] for h in ("21d", "5d", "1d") if row[f"trend_{h}"] != "INSUFFICIENT_DATA"), "INSUFFICIENT_DATA"), axis=1
    )
    all_prices["persistence_class"] = all_prices.apply(
        lambda row: "PERSISTENT_UPWARD" if row.trend_5d == row.trend_21d == "UPWARD"
        else "PERSISTENT_DOWNWARD" if row.trend_5d == row.trend_21d == "DOWNWARD"
        else "PERSISTENT_SIDEWAYS" if row.trend_5d == row.trend_21d == "SIDEWAYS"
        else "REVERSAL" if {row.trend_1d, row.trend_21d} == {"UPWARD", "DOWNWARD"}
        else "MIXED_OR_TRANSITION", axis=1
    )
    all_prices["volatility_regime"] = all_prices["realised_vol_20d_pct"].map(
        lambda value: "NOT_AVAILABLE" if pd.isna(value) else "LOW" if value <= 15 else "NORMAL" if value <= 25 else "HIGH" if value <= 40 else "EXTREME"
    )
    vix = index_df[index_df.symbol == "INDIA VIX"][["trade_date", "close_price"]].rename(columns={"close_price": "india_vix"})
    vix["trade_date"] = pd.to_datetime(vix["trade_date"])
    all_prices = all_prices.merge(vix, on="trade_date", how="left")
    bands = policy["vix_level_bands"]
    all_prices["vix_regime"] = all_prices["india_vix"].map(
        lambda value: "NOT_AVAILABLE" if pd.isna(value) else "LOW" if value <= bands["low_max"] else "NORMAL" if value <= bands["normal_max"] else "HIGH" if value <= bands["high_max"] else "EXTREME"
    )
    all_prices["market_zone"] = all_prices.apply(
        lambda row: "INSUFFICIENT_DATA" if row.primary_trend == "INSUFFICIENT_DATA"
        else f"{row.primary_trend}_{'HIGH_VOL' if row.volatility_regime in {'HIGH','EXTREME'} or row.vix_regime in {'HIGH','EXTREME'} else 'LOW_NORMAL_VOL'}", axis=1
    )
    rows = []
    for record in all_prices.itertuples(index=False):
        rows.append((
            record.trade_date.date(), record.instrument_type, record.symbol, POLICY_VERSION,
            None if pd.isna(record.close_price) else float(record.close_price),
            *[None if pd.isna(getattr(record, f"return_{h}_pct")) else float(getattr(record, f"return_{h}_pct")) for h in ("1d", "5d", "21d", "63d")],
            record.trend_1d, record.trend_5d, record.trend_21d, record.trend_63d,
            record.primary_trend, record.persistence_class,
            None if pd.isna(record.realised_vol_20d_pct) else float(record.realised_vol_20d_pct),
            record.volatility_regime, None if pd.isna(record.india_vix) else float(record.india_vix),
            record.vix_regime, record.market_zone,
            "OK" if record.primary_trend != "INSUFFICIENT_DATA" else "WARMUP_INCOMPLETE",
            latest_batch if record.instrument_type == "STOCK" else None,
        ))
    cur.executemany(
        """
        INSERT INTO strategy_eval.market_regime_daily (
          trade_date,instrument_type,symbol,policy_version,close_price,return_1d_pct,return_5d_pct,
          return_21d_pct,return_63d_pct,trend_1d,trend_5d,trend_21d,trend_63d,primary_trend,
          persistence_class,realised_vol_20d_pct,volatility_regime,india_vix,vix_regime,market_zone,
          data_quality_flag,source_batch_run_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (trade_date,instrument_type,symbol,policy_version) DO UPDATE SET
          close_price=EXCLUDED.close_price,return_1d_pct=EXCLUDED.return_1d_pct,return_5d_pct=EXCLUDED.return_5d_pct,
          return_21d_pct=EXCLUDED.return_21d_pct,return_63d_pct=EXCLUDED.return_63d_pct,
          trend_1d=EXCLUDED.trend_1d,trend_5d=EXCLUDED.trend_5d,trend_21d=EXCLUDED.trend_21d,
          trend_63d=EXCLUDED.trend_63d,primary_trend=EXCLUDED.primary_trend,
          persistence_class=EXCLUDED.persistence_class,realised_vol_20d_pct=EXCLUDED.realised_vol_20d_pct,
          volatility_regime=EXCLUDED.volatility_regime,india_vix=EXCLUDED.india_vix,vix_regime=EXCLUDED.vix_regime,
          market_zone=EXCLUDED.market_zone,data_quality_flag=EXCLUDED.data_quality_flag,
          source_batch_run_id=EXCLUDED.source_batch_run_id,calculated_at=NOW()
        """, rows,
    )
    return len(rows)


def evaluate_latest_runs(cur) -> int:
    runs = fetch_frame(
        cur,
        """
        WITH latest AS (
          SELECT batch_run_id FROM nse_app.batch_run_audit
          WHERE batch_name='backtesting_precompute' AND published_flag
          ORDER BY data_as_of_date DESC,generated_at DESC LIMIT 1
        )
        SELECT r.backtest_run_id,r.batch_run_id,r.universe_mode,r.capital_mode,r.summary_json,
               v.config_json,v.assumptions_json
        FROM nse_app.backtest_run r JOIN latest USING(batch_run_id)
        JOIN nse_app.backtest_strategy_version v USING(strategy_version_id)
        WHERE r.universe_mode='nifty_100'
        ORDER BY r.backtest_run_id
        """,
    )
    for row in runs.itertuples(index=False):
        summary = row.summary_json if isinstance(row.summary_json, dict) else json.loads(row.summary_json)
        config = row.config_json if isinstance(row.config_json, dict) else json.loads(row.config_json)
        assumptions = row.assumptions_json if isinstance(row.assumptions_json, dict) else json.loads(row.assumptions_json)
        decision = evaluate_rankability(
            config, assumptions, capital_mode=row.capital_mode, universe_mode=row.universe_mode,
            closed_trades=int(summary.get("totalClosedTrades") or 0), open_positions_included="openPositions" in summary,
        )
        cur.execute(
            """
            INSERT INTO strategy_eval.run_evaluation (
              backtest_run_id,policy_version,result_type,rankability_status,rating,quality_score,
              evidence_multiplier,revenue_capacity_score,validation_status,validation_json,
              good_when_json,avoid_when_json,watch_json,limitation_json,evaluator_version)
            VALUES (%s,%s,%s,%s,%s,NULL,NULL,NULL,%s,%s::jsonb,'[]','[]','[]',%s::jsonb,%s)
            ON CONFLICT (backtest_run_id,policy_version) DO UPDATE SET
              result_type=EXCLUDED.result_type,rankability_status=EXCLUDED.rankability_status,
              rating=EXCLUDED.rating,quality_score=NULL,evidence_multiplier=NULL,revenue_capacity_score=NULL,
              validation_status=EXCLUDED.validation_status,validation_json=EXCLUDED.validation_json,
              limitation_json=EXCLUDED.limitation_json,evaluated_at=NOW(),evaluator_version=EXCLUDED.evaluator_version
            """,
            (row.backtest_run_id, POLICY_VERSION, decision.result_type, decision.rankability_status,
             decision.rating, decision.validation_status, as_json(decision.gates), as_json(decision.limitations), EVALUATOR_VERSION),
        )
    cur.execute(
        """
        INSERT INTO strategy_eval.trade_context_snapshot (
          trade_log_id,stock_regime_date,stock_primary_trend,stock_market_zone,nifty_primary_trend,
          nifty_market_zone,india_vix,vix_regime,event_ids,context_json)
        SELECT t.trade_log_id,t.entry_date,s.primary_trend,s.market_zone,n.primary_trend,n.market_zone,
               n.india_vix,n.vix_regime,
               COALESCE((SELECT jsonb_agg(e.event_id ORDER BY e.event_id) FROM strategy_eval.market_event e WHERE e.anchor_session=t.entry_date),'[]'::jsonb),
               jsonb_build_object('policy_version',%s::text,'retrospective_events_not_entry_features',true)
        FROM nse_app.backtest_trade_log t
        JOIN strategy_eval.run_evaluation re ON re.policy_version=%s
        JOIN nse_app.backtest_run r ON r.backtest_run_id=re.backtest_run_id
          AND r.batch_run_id=t.batch_run_id AND r.strategy_version_id=t.strategy_version_id AND r.scenario_key=t.scenario_key
        LEFT JOIN strategy_eval.market_regime_daily s ON s.trade_date=t.entry_date AND s.instrument_type='STOCK'
          AND s.symbol=t.symbol AND s.policy_version=%s
        LEFT JOIN strategy_eval.market_regime_daily n ON n.trade_date=t.entry_date AND n.instrument_type='INDEX'
          AND n.symbol='NIFTY 50' AND n.policy_version=%s
        ON CONFLICT (trade_log_id) DO UPDATE SET
          stock_regime_date=EXCLUDED.stock_regime_date,stock_primary_trend=EXCLUDED.stock_primary_trend,
          stock_market_zone=EXCLUDED.stock_market_zone,nifty_primary_trend=EXCLUDED.nifty_primary_trend,
          nifty_market_zone=EXCLUDED.nifty_market_zone,india_vix=EXCLUDED.india_vix,
          vix_regime=EXCLUDED.vix_regime,event_ids=EXCLUDED.event_ids,context_json=EXCLUDED.context_json
        """, (POLICY_VERSION, POLICY_VERSION, POLICY_VERSION, POLICY_VERSION),
    )
    cur.execute("DELETE FROM strategy_eval.slice_metric WHERE evaluation_id IN (SELECT evaluation_id FROM strategy_eval.run_evaluation WHERE policy_version=%s)", (POLICY_VERSION,))
    cur.execute(
        """
        INSERT INTO strategy_eval.slice_metric (evaluation_id,slice_type,slice_key,sample_size,metrics_json,suitability)
        SELECT re.evaluation_id,x.slice_type,x.slice_key,COUNT(*)::int,
               jsonb_build_object('avg_return_pct',ROUND(AVG(t.return_pct),4),'median_return_pct',ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY t.return_pct)::numeric,4),
                 'win_rate_pct',ROUND(100.0*AVG(CASE WHEN COALESCE(t.after_tax_net_pnl,t.net_pnl)>0 THEN 1 ELSE 0 END),2),
                 'after_tax_net_pnl',ROUND(SUM(COALESCE(t.after_tax_net_pnl,t.net_pnl)),2)),
               CASE WHEN re.result_type='OPPORTUNITY_SCAN' THEN 'UNKNOWN'
                    WHEN COUNT(*)<20 THEN 'UNKNOWN'
                    WHEN AVG(t.return_pct)>0 THEN 'GOOD' ELSE 'AVOID' END
        FROM strategy_eval.run_evaluation re
        JOIN nse_app.backtest_run r ON r.backtest_run_id=re.backtest_run_id
        JOIN nse_app.backtest_trade_log t ON t.batch_run_id=r.batch_run_id AND t.strategy_version_id=r.strategy_version_id AND t.scenario_key=r.scenario_key
        JOIN strategy_eval.trade_context_snapshot c ON c.trade_log_id=t.trade_log_id
        CROSS JOIN LATERAL (VALUES
          ('NIFTY_TREND',COALESCE(c.nifty_primary_trend,'UNKNOWN')),
          ('STOCK_TREND',COALESCE(c.stock_primary_trend,'UNKNOWN')),
          ('STOCK_NIFTY_MATRIX',COALESCE(c.stock_primary_trend,'UNKNOWN')||' / '||COALESCE(c.nifty_primary_trend,'UNKNOWN')),
          ('VIX_REGIME',COALESCE(c.vix_regime,'UNKNOWN'))
        ) x(slice_type,slice_key)
        WHERE re.policy_version=%s AND t.trade_status='closed' AND t.return_pct IS NOT NULL
        GROUP BY re.evaluation_id,x.slice_type,x.slice_key
        """, (POLICY_VERSION,),
    )
    cur.execute(
        """
        UPDATE strategy_eval.run_evaluation re SET
          good_when_json=COALESCE((SELECT jsonb_agg(jsonb_build_object('context',slice_type,'value',slice_key,'sample_size',sample_size,'metrics',metrics_json) ORDER BY sample_size DESC) FROM strategy_eval.slice_metric sm WHERE sm.evaluation_id=re.evaluation_id AND sm.suitability='GOOD'),'[]'::jsonb),
          avoid_when_json=COALESCE((SELECT jsonb_agg(jsonb_build_object('context',slice_type,'value',slice_key,'sample_size',sample_size,'metrics',metrics_json) ORDER BY sample_size DESC) FROM strategy_eval.slice_metric sm WHERE sm.evaluation_id=re.evaluation_id AND sm.suitability='AVOID'),'[]'::jsonb),
          watch_json=COALESCE((SELECT jsonb_agg(jsonb_build_object('context',slice_type,'value',slice_key,'sample_size',sample_size,'metrics',metrics_json) ORDER BY sample_size DESC) FROM strategy_eval.slice_metric sm WHERE sm.evaluation_id=re.evaluation_id AND sm.suitability IN ('WATCH','UNKNOWN')),'[]'::jsonb)
        WHERE re.policy_version=%s
        """, (POLICY_VERSION,),
    )
    return len(runs)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--rules", type=Path, required=True)
    parser.add_argument("--policy", type=Path, default=PROJECT_ROOT / "config/evaluation/strategy_evaluation_roe_v1.json")
    parser.add_argument("--schema-sql", type=Path, default=MONOREPO_ROOT / "db/sql/020_strategy_evaluation_roe.sql")
    parser.add_argument("--dry-run", action="store_true")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if not args.database_url:
        raise SystemExit("--database-url or DATABASE_URL is required")
    for path in (args.workbook, args.rules, args.policy, args.schema_sql):
        if not path.is_file():
            raise SystemExit(f"Required file not found: {path}")
    with psycopg.connect(args.database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(args.schema_sql.read_text(encoding="utf-8"))
            policy = insert_policy(cur, args.policy, args.rules)
            metrics = ingest_workbook(cur, args.workbook)
            metrics["regime_rows"] = calculate_regimes(cur, policy)
            metrics["evaluated_runs"] = evaluate_latest_runs(cur)
            metrics["policy_version"] = POLICY_VERSION
            metrics["workbook_sha256"] = sha256(args.workbook)
            metrics["rules_sha256"] = sha256(args.rules)
            if args.dry_run:
                conn.rollback()
                metrics["committed"] = False
            else:
                conn.commit()
                metrics["committed"] = True
    print(json.dumps(metrics, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
