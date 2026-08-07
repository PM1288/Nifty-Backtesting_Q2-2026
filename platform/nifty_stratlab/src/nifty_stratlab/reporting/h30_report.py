"""Uniform H30 CSV, Parquet, JSON, Markdown, Excel and chart outputs."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

from .month_density_charts import render_strategy_charts


def flatten_observation(row: dict[str, Any]) -> dict[str, Any]:
    flat = {k: v for k, v in row.items() if k not in {"checkpoints", "max_close_economics", "swing_d5_economics", "opportunity_bands", "rank_blockers"}}
    flat["rank_blockers"] = "|".join(row.get("rank_blockers", []))
    for prefix, payload in (("h30", row.get("max_close_economics") or {}), ("d5", row.get("swing_d5_economics") or {})):
        for key, value in payload.items():
            if not isinstance(value, dict): flat[f"{prefix}_{key}"] = value
    return flat


def _intraday_max_net_upside(trade: dict[str, Any]) -> float | None:
    d0 = next((row for row in trade.get("path_checkpoints", []) if row.get("stage") == "D0_CLOSE"), None)
    if not d0 or d0.get("mfe_pct") is None:
        return None
    cost_bps = float((trade.get("policy") or {}).get("intraday_round_trip_cost_bps", 8))
    return round(float(d0["mfe_pct"]) - cost_bps / 100.0, 8)


def write_h30_report(output_dir: Path, strategy_id: str, trades: list[dict[str, Any]], ranking: dict[str, Any]) -> tuple[list[Path], pd.DataFrame]:
    observations = [r["h30_observation"] for r in trades if r.get("h30_observation")]
    flat = pd.DataFrame([flatten_observation(r) for r in observations])
    checkpoints = pd.DataFrame([{**c, "symbol": r["symbol"], "entry_date": r["entry_date"]} for r in observations for c in r.get("checkpoints", [])])
    flat.to_csv(output_dir / "h30_observations.csv", index=False)
    checkpoints.to_csv(output_dir / "h30_checkpoints.csv", index=False)
    flat.to_parquet(output_dir / "h30_observations.parquet", index=False)
    checkpoints.to_parquet(output_dir / "h30_checkpoints.parquet", index=False)
    (output_dir / "h30_ranking.json").write_text(json.dumps(ranking, indent=2, sort_keys=True) + "\n")
    (output_dir / "h30_summary.md").write_text(
        "# H30 maximum-close opportunity\n\n"
        "> Hypothetical hindsight entry-quality evidence. It is not realised P&L and never changes execution exits.\n\n"
        f"- Status: `{ranking['status']}`\n- Diagnostic score: `{ranking['diagnostic_score']}`\n"
        f"- Mature observations: `{ranking['summary']['mature_entry_count']}`\n"
        f"- Hard-gate blockers: `{', '.join(ranking['hard_gate_blockers']) or 'none'}`\n"
    )
    chart = pd.DataFrame({
        "entry_path_id": flat.get("entry_path_id", pd.Series(dtype=str)), "entry_date": flat.get("entry_date", pd.Series(dtype=str)),
        "coverage_status": flat.get("coverage_status", pd.Series(dtype=str)),
        "intraday_max_net_upside_pct": [_intraday_max_net_upside(t) for t in trades if t.get("h30_observation")],
        "swing_d5_max_close_after_tax_upside_pct": flat.get("d5_after_tax_upside_pct", pd.Series(dtype=float)),
        "h30_max_close_after_tax_upside_pct": flat.get("h30_after_tax_upside_pct", pd.Series(dtype=float)),
    })
    charts = render_strategy_charts(chart, strategy_id, output_dir)
    wb = Workbook(); ws = wb.active; ws.title = "00_READ_ME"
    ws.append(["H30 evaluation report"]); ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Semantics", "Hypothetical maximum official close; not realised P&L; no early exit."])
    ws.append(["Ranking status", ranking["status"]]); ws.append(["Blockers", ", ".join(ranking["hard_gate_blockers"])])
    for title, frame in (("06_H30_OBSERVATIONS", flat), ("07_H30_CHECKPOINTS", checkpoints), ("08_H30_RANKING", pd.DataFrame([ranking]))):
        sheet = wb.create_sheet(title)
        if frame.empty: sheet.append(["No rows"]); continue
        sheet.append(list(frame.columns))
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1E3A5F")
        for values in frame.fillna("").itertuples(index=False, name=None):
            sheet.append([json.dumps(v, sort_keys=True) if isinstance(v, (dict, list)) else v for v in values])
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    chart_sheet = wb.create_sheet("09_H30_CHARTS")
    chart_sheet.add_image(Image(str(charts["intraday_png"])), "A1"); chart_sheet.add_image(Image(str(charts["swing_long_png"])), "A32")
    wb.save(output_dir / "strategy_evaluation.xlsx")
    return [output_dir / name for name in ("h30_observations.csv", "h30_checkpoints.csv", "h30_observations.parquet", "h30_checkpoints.parquet", "h30_ranking.json", "h30_summary.md", "strategy_evaluation.xlsx")] + list(charts.values()), flat
